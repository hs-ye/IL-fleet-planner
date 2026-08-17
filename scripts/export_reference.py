"""One-time export: IL Fleet Planner.xlsx -> reference-data JSON files for the web app.

Splits the reference data into three flat files so each can be inspected/edited on its own:
  public/reference-data/ships.json      (bare array of ships)
  public/reference-data/modules.json    (bare array of modules)
  public/reference-data/templates.json  (bare array of templates)

Run:  uv run --with openpyxl python scripts/export_reference.py
"""
import json
import os
from openpyxl import load_workbook

EXCEL = r"C:\Users\yehan\Dropbox\Analysis\IL Fleet Planner.xlsx"
OUT_DIR = r"C:\GitRepos\IL-fleet-planner\public\reference-data"


def clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v != "" else None
    return v


def num(v):
    if v is None or v == "":
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


wb = load_workbook(EXCEL, data_only=True)

# ---- Ships (10 cols: Ship, Base, Variant, Class, CP, Row, Corv, Fighter, Size, Notes) ----
ships = []
name_to_key = {}
used_keys = set()
for r in wb["Ships"].iter_rows(min_row=2):
    name = clean(r[0].value)
    if not name:
        continue
    base = clean(r[1].value)
    variant = clean(r[2].value)
    key = f"{base}:{variant}" if (base and variant) else (base or name)
    if key in used_keys:
        i = 2
        while f"{key}~{i}" in used_keys:
            i += 1
        key = f"{key}~{i}"
    used_keys.add(key)
    name_to_key[name] = key
    ships.append({
        "key": key,
        "name": name,
        "base": base,
        "variant": variant,
        "class": clean(r[3].value),
        "cpCost": num(r[4].value),
        "defaultRow": clean(r[5].value),
        "fixedCorvSlots": num(r[6].value),
        "fixedFighterSlots": num(r[7].value),
        "size": clean(r[8].value),
        "notes": clean(r[9].value),
    })

# ---- Modules (7 cols: Ship, Slot, Module Name, Corv, Fighter, Hangar Type, Effect) ----
mod_by_name = {}
modules = []
for r in wb["Modules"].iter_rows(min_row=2):
    name = clean(r[2].value)
    if not name:
        continue
    ship_name = clean(r[0].value)
    ship_key = name_to_key.get(ship_name, ship_name)
    corv = num(r[3].value) or 0
    ftr = num(r[4].value) or 0
    mod_by_name[name] = (corv, ftr)
    modules.append({
        "shipKey": ship_key,
        "slot": clean(r[1].value),
        "name": name,
        "corvCapacity": corv,
        "fighterCapacity": ftr,
        "hangarSize": clean(r[5].value),
        "effect": clean(r[6].value),
    })

# ---- Templates (12 cols: Name, Base Ship, M..F(3-9), Corv(10), Fighter(11), Notes(12)) ----
SLOTS = ["M", "A", "B", "C", "D", "E", "F"]
templates = []
for r in wb["Templates"].iter_rows(min_row=2):
    name = clean(r[0].value)
    if not name:
        continue
    base_ship = clean(r[1].value)
    slots = {}
    corv_total = 0
    ftr_total = 0
    for i, s in enumerate(SLOTS):
        mod_name = clean(r[2 + i].value)
        if mod_name:
            slots[s] = mod_name
            c, f = mod_by_name.get(mod_name, (0, 0))
            corv_total += c
            ftr_total += f
    templates.append({
        "key": name,
        "name": name,
        "baseShipKey": name_to_key.get(base_ship, base_ship),
        "slots": slots,
        "corvSlots": corv_total,
        "fighterSlots": ftr_total,
    })

os.makedirs(OUT_DIR, exist_ok=True)

def write_json(filename, payload):
    path = os.path.join(OUT_DIR, filename)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"  {path}  ({len(payload)} entries)")

write_json("ships.json", ships)
write_json("modules.json", modules)
write_json("templates.json", templates)

empty_base = [s["name"] for s in ships if not s["base"]]
print(f"ships with empty base (key=name): {len(empty_base)}")
